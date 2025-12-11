from datetime import date
from typing import List, Optional
from openpyxl import load_workbook
from app.club_members.schemas import ClubMemberSchema


class ClubMembers:
    def __init__(self, filename: str):
        self.filename = filename
        self.members: List[ClubMemberSchema] = []

        wb = load_workbook(filename, data_only=True)
        sheet = wb["Members"] 
        headers = [cell.value for cell in sheet[1]]

        idx_club_id = headers.index("Club ID")
        idx_name = headers.index("Name")
        idx_spreadsheet = headers.index("Spreadsheet Key")
        idx_sync_date = headers.index("Sync Date") if "Sync Date" in headers else None

        for row in sheet.iter_rows(min_row=2, values_only=True):
            raw_club_id = row[idx_club_id]
            if raw_club_id is None:
                continue 

            club_id = int(raw_club_id)
            name = str(row[idx_name])
            spreadsheet_key = str(row[idx_spreadsheet])

            sync_date = None
            if idx_sync_date is not None and row[idx_sync_date] is not None:
                if isinstance(row[idx_sync_date], date):
                    sync_date = row[idx_sync_date]
                else:
                    sync_date = date.fromisoformat(str(row[idx_sync_date]))

            member = ClubMemberSchema(
                club_id=club_id,
                name=name,
                spreadsheet_key=spreadsheet_key,
                sync_date=sync_date,
            )
            self.members.append(member)

        wb.close()

    def save(self, filename: Optional[str] = None):
        filename = filename or self.filename

        wb = load_workbook(filename)
        ws = wb["Members"]
        for i, m in enumerate(self.members, start=2):
            ws[f"D{i}"] = m.sync_date
        wb.save(filename)
